# -*- coding: utf-8 -*-

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import os
import shutil
import winreg
from datetime import datetime
import getpass
import socket
import re

# ========== Excel (untuk weekly collect & grafik) ==========
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.chart import LineChart, BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except Exception:
    OPENPYXL_OK = False


class StorageMonitorApp:
    def __init__(self, root):
        self.root = root
        self.timer_seconds = 5  # auto-close setelah 5 detik

        # Nama bulan (untuk nama file TXT)
        self.month_names = {
            1: "januari", 2: "februari", 3: "maret", 4: "april",
            5: "mei", 6: "juni", 7: "juli", 8: "agustus",
            9: "september", 10: "oktober", 11: "november", 12: "desember"
        }

        # Abaikan tombol X
        self.root.protocol("WM_DELETE_WINDOW", self.on_close_attempt)

        self.setup_ui()
        self.update_storage_info()
        self.load_installed_applications()
        self.start_timer()

    def on_close_attempt(self):
        # Klik X diabaikan
        pass

    # ================= UI =================
    def setup_ui(self):
        self.root.title("System Info - Storage & Apps")
        self.root.geometry("700x600")
        self.root.configure(bg="#1e1e1e")
        self.root.resizable(True, True)

        style = ttk.Style()
        for theme in ("alt", "clam", "default"):
            try:
                style.theme_use(theme)
                break
            except Exception:
                continue
        style.configure("Title.TLabel", font=("Arial", 16, "bold"),
                        foreground="#64B5F6", background="#1e1e1e")
        style.configure("Info.TLabel", font=("Arial", 10),
                        foreground="#B0B0B0", background="#1e1e1e")
        style.configure("Timer.TLabel", font=("Arial", 14, "bold"),
                        foreground="#FF7043", background="#1e1e1e")
        style.configure("TFrame", background="#1e1e1e")
        style.configure("TLabelframe", background="#1e1e1e", foreground="#E0E0E0")
        style.configure("TLabelframe.Label", background="#1e1e1e", foreground="#64B5F6")
        style.configure("TButton", background="#2d2d2d", foreground="#E0E0E0")
        style.map("TButton", background=[("active", "#404040")])
        style.configure("TProgressbar", background="#64B5F6", troughcolor="#2d2d2d")

        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        computer_name = socket.gethostname().upper()
        ttk.Label(main_frame, text=f"Computer: {computer_name}",
                  style="Title.TLabel").pack(pady=(0, 20))

        self.timer_label = ttk.Label(main_frame,
                                     text=f"Auto-close in: {int(self.timer_seconds)}s",
                                     style="Timer.TLabel")
        self.timer_label.pack(pady=(0, 10))

        # Storage info
        storage_frame = ttk.LabelFrame(main_frame, text="Drive C: Storage Info", padding="15")
        storage_frame.pack(fill=tk.X, pady=(0, 20))

        self.storage_info_label = ttk.Label(storage_frame,
                                            text="Loading storage information...",
                                            style="Info.TLabel")
        self.storage_info_label.pack(anchor=tk.W)

        self.storage_progress = ttk.Progressbar(storage_frame, length=400, mode="determinate")
        self.storage_progress.pack(fill=tk.X, pady=(10, 0))

        # Installed apps (kode lama dipertahankan)
        apps_frame = ttk.LabelFrame(main_frame, text="Installed Applications", padding="15")
        apps_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))

        self.apps_text = scrolledtext.ScrolledText(
            apps_frame, height=12, width=70, font=("Consolas", 9),
            bg="#2d2d2d", fg="#E0E0E0", insertbackground="#E0E0E0",
            selectbackground="#404040", wrap=tk.WORD
        )
        self.apps_text.pack(fill=tk.BOTH, expand=True)

        # Buttons
        buttons_frame = ttk.Frame(main_frame)
        buttons_frame.pack(fill=tk.X, pady=(0, 20))

        ttk.Button(buttons_frame, text="Export Info (TXT)", command=self.export_info)\
            .pack(side=tk.RIGHT, padx=(10, 0))

        # Manual: FORCE tulis minggu ini
        ttk.Button(buttons_frame, text="Update Excel Sekarang",
                   command=lambda: self.update_excel_weekly(force=True))\
            .pack(side=tk.RIGHT, padx=(10, 0))

        # Manual: Bangun Excel dari TXT lama
        ttk.Button(buttons_frame, text="Build Excel dari TXT",
                   command=self.rebuild_excel_from_txt)\
            .pack(side=tk.RIGHT, padx=(10, 0))

        ttk.Label(main_frame, text="© JEMI OKTAVIAN", style="Info.TLabel")\
            .pack(side=tk.BOTTOM)

    # ============== Utility Path ==============
    def _network_root(self):
        """Root share utama di server."""
        return r"\\192.168.9.23\TXT STORAGE MONITOR"

    def _get_base_dir(self):
        """
        Base folder penyimpanan (utama ke network):
        \\192.168.9.23\TXT STORAGE MONITOR\<PC_NAME>
        Fallback: C:\TXT STORAGE MONITOR\<PC_NAME> -> Documents\TXT_STORAGE_MONITOR\<PC_NAME>
        """
        pc_name = os.environ.get("COMPUTERNAME") or socket.gethostname()
        # 1) Network utama
        net_base = os.path.join(self._network_root(), pc_name)
        try:
            os.makedirs(net_base, exist_ok=True)
            return net_base
        except Exception:
            pass

        # 2) Fallback ke lokal C:
        local_base = os.path.join(r"C:\TXT STORAGE MONITOR", pc_name)
        try:
            os.makedirs(local_base, exist_ok=True)
            return local_base
        except Exception:
            pass

        # 3) Fallback terakhir: Documents
        docs_base = os.path.join(os.path.expanduser("~"), "Documents", "TXT_STORAGE_MONITOR", pc_name)
        os.makedirs(docs_base, exist_ok=True)
        return docs_base

    def _resolve_export_path(self, filename):
        """
        Lokasi file TXT (mengikuti kebijakan yang sama dengan _get_base_dir()).
        """
        base = self._get_base_dir()
        return os.path.join(base, filename)

    # ============== LOGIC ==============
    def update_storage_info(self):
        try:
            total, used, free = shutil.disk_usage("C:\\")
            total_gb = total / (1024**3)
            used_gb = used / (1024**3)
            free_gb = free / (1024**3)
            used_percent = round((used_gb / total_gb) * 100, 2)
            free_percent = round(100.0 - used_percent, 2)

            info_text = (
                f"Total Storage: {total_gb:.2f} GB (100%)\n"
                f"Used Storage:  {used_gb:.2f} GB ({used_percent:.2f}%)\n"
                f"Free Storage:  {free_gb:.2f} GB ({free_percent:.2f}%)"
            )
            self.storage_info_label.config(text=info_text)
            try:
                self.storage_progress["value"] = used_percent
            except Exception:
                pass

            self.storage_data = {
                "total_gb": round(total_gb, 2),
                "used_gb": round(used_gb, 2),
                "free_gb": round(free_gb, 2),
                "used_percent": used_percent,
                "free_percent": free_percent
            }
        except Exception as e:
            self.storage_info_label.config(text=f"Error reading storage info: {str(e)}")

    def load_installed_applications(self):
        # Kode lama dipertahankan
        try:
            self.apps_text.insert(tk.END, "DisplayName\n")
            self.apps_text.insert(tk.END, "-" * 50 + "\n")

            applications = []
            registry_paths = [
                (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall"),
                (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\WOW6432Node\Microsoft\Windows\CurrentVersion\Uninstall"),
                (winreg.HKEY_CURRENT_USER, r"SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall")
            ]

            for hkey, path in registry_paths:
                try:
                    with winreg.OpenKey(hkey, path) as key:
                        i = 0
                        while True:
                            try:
                                subkey_name = winreg.EnumKey(key, i)
                                with winreg.OpenKey(key, subkey_name) as subkey:
                                    try:
                                        display_name = winreg.QueryValueEx(subkey, "DisplayName")[0]
                                        if display_name and display_name not in applications:
                                            applications.append(display_name)
                                    except FileNotFoundError:
                                        pass
                                i += 1
                            except OSError:
                                break
                except Exception:
                    continue

            applications.sort(key=str.lower)
            self.applications_list = applications
            for app in applications:
                self.apps_text.insert(tk.END, f"{app}\n")

            self.apps_text.config(state=tk.DISABLED)

        except Exception as e:
            self.apps_text.insert(tk.END, f"Error loading applications: {str(e)}\n")

    def start_timer(self):
        def tick():
            if self.timer_seconds > 0:
                self.timer_label.config(text=f"Auto-close in: {int(self.timer_seconds)}s")
                self.timer_seconds -= 1
                self.root.after(1000, tick)
            else:
                # Otomatis: export TXT + update Excel mingguan (tanpa force)
                self.export_info(auto=True)
                self.root.quit()
        tick()

    # ===== Excel weekly collect: STORAGE ONLY =====
    def update_excel_weekly(self, force=False):
        """Tulis 1 baris per minggu (Sabtu). Minggu 1–5 berulang. Grafik Line, Bar, dan Pie (Used_GB vs Free_GB).
           Jika force=True, tulis sekarang tanpa cek hari."""
        if not OPENPYXL_OK:
            try:
                messagebox.showwarning("Excel Module Missing",
                                       "openpyxl belum terpasang. Jalankan: pip install openpyxl")
            except Exception:
                pass
            return

        now = datetime.now()
        weekday = now.weekday()  # Mon=0 ... Sun=6
        if not force:
            if weekday == 6:   # Minggu: skip
                return
            if weekday != 5:   # Hanya Sabtu yang nulis
                return

        base_dir = self._get_base_dir()
        xlsx_path = os.path.join(base_dir, "StorageMonitor_Weekly.xlsx")

        # Open or create
        if os.path.exists(xlsx_path):
            wb = load_workbook(xlsx_path)
        else:
            wb = Workbook()

        ws = wb["Data"] if "Data" in wb.sheetnames else wb.create_sheet("Data", 0)
        chart_ws = wb["Grafik"] if "Grafik" in wb.sheetnames else wb.create_sheet("Grafik", 1)

        headers = ["Minggu", "Tanggal", "Total_GB", "Used_GB", "Free_GB",
                   "Used_%", "Free_%", "Kenaikan_GB", "Penurunan_GB",
                   "Kenaikan_%", "Penurunan_%"]
        if ws.max_row < 1 or ws["A1"].value is None:
            ws.append(headers)
            ws.freeze_panes = "A2"
            col_widths = [12, 12, 10, 10, 10, 10, 10, 13, 14, 12, 13]
            for i, w in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

        # Cegah duplikasi untuk tanggal ini
        tgl_str = now.strftime("%d/%m/%Y")
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(row=r, column=2).value) == tgl_str:
                if force:
                    try:
                        messagebox.showinfo("Info", "Baris untuk tanggal ini sudah ada.")
                    except Exception:
                        pass
                wb.save(xlsx_path)
                return

        if not hasattr(self, "storage_data"):
            wb.save(xlsx_path)
            return
        sd = self.storage_data

        # Minggu 1..5 berulang
        existing_rows = ws.max_row - 1  # tanpa header
        minggu_ke = (existing_rows % 5) + 1
        minggu_label = f"Minggu {minggu_ke}"

        # Delta dibanding baris terakhir
        prev_used_gb = ws.cell(row=ws.max_row, column=4).value if ws.max_row >= 2 else None
        prev_used_pct = ws.cell(row=ws.max_row, column=6).value if ws.max_row >= 2 else None

        kenaikan_gb = penurunan_gb = kenaikan_pct = penurunan_pct = 0.0
        try:
            if prev_used_gb is not None:
                diff_gb = round(sd["used_gb"] - float(prev_used_gb), 2)
                if diff_gb > 0:
                    kenaikan_gb = diff_gb
                elif diff_gb < 0:
                    penurunan_gb = abs(diff_gb)
        except Exception:
            pass
        try:
            if prev_used_pct is not None:
                diff_pct = round(sd["used_percent"] - float(prev_used_pct), 2)
                if diff_pct > 0:
                    kenaikan_pct = diff_pct
                elif diff_pct < 0:
                    penurunan_pct = abs(diff_pct)
        except Exception:
            pass

        # Append baris minggu ini
        ws.append([
            minggu_label, tgl_str,
            sd["total_gb"], sd["used_gb"], sd["free_gb"],
            sd["used_percent"], sd["free_percent"],
            kenaikan_gb, penurunan_gb, kenaikan_pct, penurunan_pct
        ])

        # ===== Refresh charts =====
        # Hapus chart lama
        try:
            while chart_ws._charts:
                chart_ws._charts.pop()
        except Exception:
            pass

        rows = ws.max_row
        if rows >= 2:
            cats = Reference(ws, min_col=2, min_row=2, max_row=rows)  # Tanggal

            # Line: Used vs Free (GB)
            data_gb = Reference(ws, min_col=4, max_col=5, min_row=1, max_row=rows)  # include header
            line = LineChart()
            line.title = "Used vs Free (GB) - Per Minggu (Sabtu)"
            line.y_axis.title = "GB"
            line.x_axis.title = "Tanggal"
            line.add_data(data_gb, titles_from_data=True)
            line.set_categories(cats)
            chart_ws.add_chart(line, "A2")

            # Bar: Kenaikan/Penurunan (GB)
            data_delta = Reference(ws, min_col=8, max_col=9, min_row=1, max_row=rows)
            bar = BarChart()
            bar.type = "col"
            bar.title = "Kenaikan / Penurunan (GB) Mingguan"
            bar.y_axis.title = "GB"
            bar.x_axis.title = "Tanggal"
            bar.add_data(data_delta, titles_from_data=True)
            bar.set_categories(cats)
            chart_ws.add_chart(bar, "A20")

            # Pie: Used_GB vs Free_GB (minggu terakhir)
            last_used_gb = ws.cell(row=rows, column=4).value
            last_free_gb = ws.cell(row=rows, column=5).value
            chart_ws["E2"] = "Komposisi Terakhir (GB)"
            chart_ws["E4"] = "Used_GB"; chart_ws["F4"] = last_used_gb
            chart_ws["E5"] = "Free_GB"; chart_ws["F5"] = last_free_gb

            pie = PieChart()
            labels = Reference(chart_ws, min_col=5, min_row=4, max_row=5)  # E4:E5
            data = Reference(chart_ws, min_col=6, min_row=4, max_row=5)   # F4:F5
            pie.add_data(data, titles_from_data=False)
            pie.set_categories(labels)
            pie.title = "Pie: Used vs Free (GB) - Minggu Terakhir"
            chart_ws.add_chart(pie, "E6")

        wb.save(xlsx_path)

        if force:
            try:
                messagebox.showinfo("Excel Updated", "Baris minggu ini ditulis (FORCE).")
            except Exception:
                pass

    # ===== Konversi riwayat TXT → Excel =====
    def rebuild_excel_from_txt(self):
        if not OPENPYXL_OK:
            try:
                messagebox.showwarning("Excel Module Missing",
                                       "openpyxl belum terpasang. Jalankan: pip install openpyxl")
            except Exception:
                pass
            return

        base_dir = self._get_base_dir()
        month_map = {
            "januari": 1, "februari": 2, "maret": 3, "april": 4, "mei": 5, "juni": 6,
            "juli": 7, "agustus": 8, "september": 9, "oktober": 10, "november": 11, "desember": 12
        }

        txt_files = [f for f in os.listdir(base_dir) if f.lower().endswith(".txt")]
        if not txt_files:
            try:
                messagebox.showinfo("Info", "Tidak ditemukan file TXT di folder ini.")
            except Exception:
                pass
            return

        def parse_txt(path):
            total_gb = used_gb = free_gb = used_pct = free_pct = None
            with open(path, "r", encoding="utf-8", errors="ignore") as fh:
                for line in fh:
                    low = line.strip().lower()
                    if low.startswith("total storage"):
                        m = re.search(r"([\d\.]+)\s*GB", line)
                        if m:
                            total_gb = float(m.group(1))
                    elif low.startswith("used  storage") or low.startswith("used storage"):
                        m = re.search(r"([\d\.]+)\s*GB.*?\(([\d\.]+)%\)", line)
                        if m:
                            used_gb = float(m.group(1)); used_pct = float(m.group(2))
                    elif low.startswith("free  storage") or low.startswith("free storage"):
                        m = re.search(r"([\d\.]+)\s*GB.*?\(([\d\.]+)%\)", line)
                        if m:
                            free_gb = float(m.group(1)); free_pct = float(m.group(2))
            return total_gb, used_gb, free_gb, used_pct, free_pct

        def parse_date_from_filename(fn):
            # contoh: NAMA_PC_username_19_agustus_2025.txt
            m = re.search(r"_(\d{1,2})_([a-zA-Z\-]+)_(\d{4})\.txt$", fn, flags=re.IGNORECASE)
            if not m:
                return None
            day = int(m.group(1))
            month_name = m.group(2).lower()
            month = month_map.get(month_name)
            year = int(m.group(3))
            if not month:
                return None
            return datetime(year, month, day)

        # Kumpulkan record terakhir per ISO-week (skip Minggu)
        records = {}
        for fn in txt_files:
            dt = parse_date_from_filename(fn)
            if not dt:
                continue
            if dt.weekday() == 6:  # Minggu skip
                continue
            total_gb, used_gb, free_gb, used_pct, free_pct = parse_txt(os.path.join(base_dir, fn))
            if used_gb is None or free_gb is None or total_gb is None:
                continue
            key = (dt.isocalendar().year, dt.isocalendar().week)
            prev = records.get(key)
            if (not prev) or (dt > prev["date"]):
                records[key] = {
                    "date": dt,
                    "total_gb": round(total_gb, 2),
                    "used_gb": round(used_gb, 2),
                    "free_gb": round(free_gb, 2),
                    "used_percent": round(used_pct, 2) if used_pct is not None else round(used_gb / total_gb * 100, 2),
                    "free_percent": round(free_pct, 2) if free_pct is not None else round(100 - (used_gb / total_gb * 100), 2),
                }

        if not records:
            try:
                messagebox.showinfo("Info", "Tidak ada data yang valid untuk dikonversi.")
            except Exception:
                pass
            return

        ordered = sorted(records.items(), key=lambda kv: (kv[0][0], kv[0][1]))

        xlsx_path = os.path.join(base_dir, "StorageMonitor_Weekly.xlsx")
        wb = Workbook()
        ws = wb.active; ws.title = "Data"
        chart_ws = wb.create_sheet("Grafik", 1)

        headers = ["Minggu", "Tanggal", "Total_GB", "Used_GB", "Free_GB",
                   "Used_%", "Free_%", "Kenaikan_GB", "Penurunan_GB",
                   "Kenaikan_%", "Penurunan_%"]
        ws.append(headers)
        ws.freeze_panes = "A2"
        col_widths = [12, 12, 10, 10, 10, 10, 10, 13, 14, 12, 13]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        prev_used_gb = prev_used_pct = None
        counter = 0
        for _, rec in ordered:
            counter += 1
            minggu_ke = ((counter - 1) % 5) + 1
            tgl_str = rec["date"].strftime("%d/%m/%Y")

            kenaikan_gb = penurunan_gb = kenaikan_pct = penurunan_pct = 0.0
            if prev_used_gb is not None:
                diff_gb = round(rec["used_gb"] - prev_used_gb, 2)
                if diff_gb > 0:
                    kenaikan_gb = diff_gb
                elif diff_gb < 0:
                    penurunan_gb = abs(diff_gb)
            if prev_used_pct is not None:
                diff_pct = round(rec["used_percent"] - prev_used_pct, 2)
                if diff_pct > 0:
                    kenaikan_pct = diff_pct
                elif diff_pct < 0:
                    penurunan_pct = abs(diff_pct)

            ws.append([
                f"Minggu {minggu_ke}", tgl_str,
                rec["total_gb"], rec["used_gb"], rec["free_gb"],
                rec["used_percent"], rec["free_percent"],
                kenaikan_gb, penurunan_gb, kenaikan_pct, penurunan_pct
            ])

            prev_used_gb = rec["used_gb"]
            prev_used_pct = rec["used_percent"]

        # Charts
        rows = ws.max_row
        if rows >= 2:
            cats = Reference(ws, min_col=2, min_row=2, max_row=rows)
            data_gb = Reference(ws, min_col=4, max_col=5, min_row=1, max_row=rows)
            line = LineChart(); line.title = "Used vs Free (GB) - Per Minggu"; line.y_axis.title = "GB"; line.x_axis.title = "Tanggal"
            line.add_data(data_gb, titles_from_data=True); line.set_categories(cats); chart_ws.add_chart(line, "A2")

            data_delta = Reference(ws, min_col=8, max_col=9, min_row=1, max_row=rows)
            bar = BarChart(); bar.type = "col"; bar.title = "Kenaikan/Penurunan (GB) Mingguan"; bar.y_axis.title="GB"; bar.x_axis.title="Tanggal"
            bar.add_data(data_delta, titles_from_data=True); bar.set_categories(cats); chart_ws.add_chart(bar, "A20")

            last_used_gb = ws.cell(row=rows, column=4).value
            last_free_gb = ws.cell(row=rows, column=5).value
            chart_ws["E2"] = "Komposisi Terakhir (GB)"
            chart_ws["E4"] = "Used_GB"; chart_ws["F4"] = last_used_gb
            chart_ws["E5"] = "Free_GB"; chart_ws["F5"] = last_free_gb
            pie = PieChart()
            labels = Reference(chart_ws, min_col=5, min_row=4, max_row=5)
            data = Reference(chart_ws, min_col=6, min_row=4, max_row=5)
            pie.add_data(data, titles_from_data=False); pie.set_categories(labels)
            pie.title = "Pie: Used vs Free (GB) - Minggu Terakhir"
            chart_ws.add_chart(pie, "E6")

        wb.save(xlsx_path)
        try:
            messagebox.showinfo("Sukses", f"Excel dibangun dari TXT:\n{xlsx_path}")
        except Exception:
            pass

    # ========== Export TXT + trigger update Excel ==========
    def export_info(self, auto=False):
        try:
            username = getpass.getuser()
            computer_name = socket.gethostname()
            now = datetime.now()
            month_name = self.month_names[now.month]
            filename = f"{computer_name}_{username}_{now.day}_{month_name}_{now.year}.txt"

            filepath = self._resolve_export_path(filename)

            # Build TXT (tetap sertakan daftar aplikasi)
            lines = []
            lines.append("System Storage and Applications Report")
            lines.append(f"Generated on: {now.strftime('%Y-%m-%d %H:%M:%S')}")
            lines.append(f"Computer: {computer_name.upper()}")
            lines.append(f"User: {username}")
            lines.append("=" * 60 + "\n")
            lines.append("DRIVE C: STORAGE INFORMATION")
            lines.append("-" * 30)
            if hasattr(self, "storage_data"):
                sd = self.storage_data
                lines.append(f"Total Storage: {sd['total_gb']:.2f} GB (100%)")
                lines.append(f"Used  Storage: {sd['used_gb']:.2f} GB ({sd['used_percent']:.2f}%)")
                lines.append(f"Free  Storage: {sd['free_gb']:.2f} GB ({sd['free_percent']:.2f}%)\n")

            lines.append("INSTALLED APPLICATIONS")
            lines.append("-" * 25)
            if hasattr(self, "applications_list"):
                for i, app in enumerate(self.applications_list, 1):
                    lines.append(f"{i:3d}. {app}")
                total_apps = len(self.applications_list)
            else:
                total_apps = 0
            lines.append(f"\nTotal Applications: {total_apps}\n")
            lines.append("=" * 60)
            lines.append("Report generated by Storage Monitor v1.0")
            lines.append("Copyright © JEMI OKTAVIAN")

            with open(filepath, "w", encoding="utf-8") as f:
                f.write("\n".join(lines))

            # Otomatis jalankan update Excel mingguan (tanpa force)
            self.update_excel_weekly()

            if not auto:
                messagebox.showinfo("Export Complete", f"Data exported to:\n{filepath}")
        except Exception as e:
            if not auto:
                messagebox.showerror("Export Error", f"Failed to export data: {str(e)}")


def main():
    try:
        root = tk.Tk()
        app = StorageMonitorApp(root)
        root.mainloop()
    except Exception as e:
        try:
            messagebox.showerror("Application Error", f"Failed to start application: {str(e)}")
        except Exception:
            print(f"Application Error: {e}")


if __name__ == "__main__":
    main()
